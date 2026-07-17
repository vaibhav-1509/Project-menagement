"""Excel/PDF report export - plan.md's 'SQL Views for monthly/yearly reports
-> Export to PDF/Excel' item. Built on the same TaskAssignment/FileProcessStatus
history the on-screen Reports page already reads, so the exported numbers
never disagree with what's shown in the browser.

Excel: a flat, filterable ledger - one row per (file, process stage) attempt
in the selected range, with both the Assigned and Completion timestamp so a
manager can see exactly when work started and finished.

PDF: the same ledger plus the same completions-over-time and per-stage
breakdown charts the Reports page shows on screen, rendered server-side
(matplotlib) so the PDF is a self-contained snapshot - no browser needed to
view it.
"""

import io
from datetime import date as date_cls, datetime, timedelta

import matplotlib

matplotlib.use("Agg")  # headless - no display server available on a server process
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from app.models import FileProcessStatus, FileRecord, FileStatus, ProcessType, TaskAssignment, User


def _complete_status_id(db: Session) -> int | None:
    return db.scalar(select(FileStatus.StatusID).where(FileStatus.StatusName == "Complete"))


def detail_rows(db: Session, start: date_cls, end_exclusive: date_cls, user_id: int | None) -> list[dict]:
    """One row per assignment attempt touching this range - either assigned or
    completed within it - so a file assigned before the range but finished
    inside it (or vice versa) still shows up with both of its real timestamps."""
    query = (
        select(TaskAssignment, FileRecord.FileName, ProcessType.ProcessTypeName, User.Username)
        .join(FileRecord, FileRecord.FileID == TaskAssignment.FileID)
        .join(ProcessType, ProcessType.ProcessTypeID == TaskAssignment.ProcessTypeID)
        .join(User, User.UserID == TaskAssignment.AssignedToUserID)
        .where(
            or_(
                and_(TaskAssignment.AssignedTS >= start, TaskAssignment.AssignedTS < end_exclusive),
                and_(TaskAssignment.CompletionTS >= start, TaskAssignment.CompletionTS < end_exclusive),
            )
        )
        .order_by(TaskAssignment.AssignedTS)
    )
    if user_id is not None:
        query = query.where(TaskAssignment.AssignedToUserID == user_id)

    status_names = {s.StatusID: s.StatusName for s in db.scalars(select(FileStatus)).all()}

    rows = []
    for assignment, file_name, process_type_name, username in db.execute(query).all():
        rows.append(
            {
                "file_name": file_name,
                "process_type": process_type_name,
                "assigned_to": username,
                "status": status_names.get(assignment.StatusID, "?"),
                "assigned_ts": assignment.AssignedTS,
                "completion_ts": assignment.CompletionTS,
            }
        )
    return rows


def _completions_by_day(db: Session, complete_id: int | None, start: date_cls, end_exclusive: date_cls, user_id: int | None) -> dict:
    if complete_id is None:
        return {}
    query = select(FileProcessStatus.CompletionTS, FileProcessStatus.FileID).where(
        FileProcessStatus.StatusID == complete_id,
        FileProcessStatus.CompletionTS >= start,
        FileProcessStatus.CompletionTS < end_exclusive,
    )
    if user_id is not None:
        query = query.where(FileProcessStatus.AssignedToUserID == user_id)
    counts: dict = {}
    for completion_ts, _file_id in db.execute(query).all():
        d = completion_ts.date()
        counts[d] = counts.get(d, 0) + 1
    return counts


def _completions_by_process_type(
    db: Session, complete_id: int | None, start: date_cls, end_exclusive: date_cls, user_id: int | None
) -> list[tuple[str, int]]:
    if complete_id is None:
        return []
    query = (
        select(ProcessType.ProcessTypeName, FileProcessStatus.FileID)
        .join(ProcessType, ProcessType.ProcessTypeID == FileProcessStatus.ProcessTypeID)
        .where(
            FileProcessStatus.StatusID == complete_id,
            FileProcessStatus.CompletionTS >= start,
            FileProcessStatus.CompletionTS < end_exclusive,
        )
        .order_by(ProcessType.SortOrder)
    )
    if user_id is not None:
        query = query.where(FileProcessStatus.AssignedToUserID == user_id)
    counts: dict[str, int] = {}
    for process_type_name, _file_id in db.execute(query).all():
        counts[process_type_name] = counts.get(process_type_name, 0) + 1
    return list(counts.items())


def _time_series(db: Session, start: date_cls, end_inclusive: date_cls, user_id: int | None) -> tuple[list[str], list[int]]:
    """One labeled bucket per day across the whole [start, end_inclusive]
    range - the same generic daily shape the Report View uses
    (app/routers/reports.py::_daily_series), so the PDF's chart matches."""
    complete_id = _complete_status_id(db)
    end_exclusive = end_inclusive + timedelta(days=1)
    day_counts = _completions_by_day(db, complete_id, start, end_exclusive, user_id)

    days = (end_inclusive - start).days + 1
    labels = [(start + timedelta(days=i)).strftime("%b %d") for i in range(days)]
    values = [day_counts.get(start + timedelta(days=i), 0) for i in range(days)]
    return labels, values


def _scope_label(user_id: int | None, username: str | None) -> str:
    return f"User: {username}" if user_id is not None else "Scope: whole team"


def _range_label(start: date_cls, end_inclusive: date_cls) -> str:
    return start.isoformat() if start == end_inclusive else f"{start.isoformat()}_to_{end_inclusive.isoformat()}"


def build_excel_report(db: Session, user_id: int | None, username: str | None, start: date_cls, end_inclusive: date_cls) -> tuple[bytes, str]:
    end_exclusive = end_inclusive + timedelta(days=1)
    label = _range_label(start, end_inclusive)
    rows = detail_rows(db, start, end_exclusive, user_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Completions Report"

    ws["A1"] = "Project Management Tool - Completions Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    ws["A3"] = f"Range: {label}"
    ws["A4"] = _scope_label(user_id, username)

    header_row = 6
    headers = ["File Name", "Process Type", "Assigned To", "Status", "Assigned Timestamp", "Completion Timestamp"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True)

    for i, row in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=row["file_name"])
        ws.cell(row=i, column=2, value=row["process_type"])
        ws.cell(row=i, column=3, value=row["assigned_to"])
        ws.cell(row=i, column=4, value=row["status"])
        ws.cell(row=i, column=5, value=row["assigned_ts"].strftime("%Y-%m-%d %H:%M:%S") if row["assigned_ts"] else "")
        ws.cell(
            row=i, column=6, value=row["completion_ts"].strftime("%Y-%m-%d %H:%M:%S") if row["completion_ts"] else ""
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), label


def _bar_chart_png(labels: list[str], values: list[int], title: str) -> bytes | None:
    if not labels:
        return None
    fig, ax = plt.subplots(figsize=(6.5, 2.8))
    ax.bar(labels, values, color="#4C72B0")
    ax.set_title(title, fontsize=11)
    ax.tick_params(axis="x", labelrotation=45, labelsize=7)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _pie_chart_png(breakdown: list[tuple[str, int]], title: str) -> bytes | None:
    if not breakdown:
        return None
    labels = [name for name, _ in breakdown]
    values = [count for _, count in breakdown]
    fig, ax = plt.subplots(figsize=(3.6, 3.2))
    if sum(values) == 0:
        ax.text(0.5, 0.5, "No completions in range", ha="center", va="center")
        ax.axis("off")
    else:
        ax.pie(values, labels=labels, autopct="%1.0f%%", textprops={"fontsize": 8})
    ax.set_title(title, fontsize=11)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def build_pdf_report(db: Session, user_id: int | None, username: str | None, start: date_cls, end_inclusive: date_cls) -> tuple[bytes, str]:
    end_exclusive = end_inclusive + timedelta(days=1)
    label = _range_label(start, end_inclusive)
    rows = detail_rows(db, start, end_exclusive, user_id)
    complete_id = _complete_status_id(db)
    process_breakdown = _completions_by_process_type(db, complete_id, start, end_exclusive, user_id)
    total_completions = sum(count for _, count in process_breakdown)
    series_labels, series_values = _time_series(db, start, end_inclusive, user_id)

    styles = getSampleStyleSheet()
    story = [
        Paragraph("Project Management Tool - Completions Report", styles["Title"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", styles["Normal"]),
        Paragraph(f"Range: {label}", styles["Normal"]),
        Paragraph(_scope_label(user_id, username), styles["Normal"]),
        Spacer(1, 0.25 * inch),
        Paragraph(f"Total stage completions in range: {total_completions}", styles["Heading3"]),
        Spacer(1, 0.15 * inch),
    ]

    bar_png = _bar_chart_png(series_labels, series_values, f"Completions - {label}")
    if bar_png:
        story.append(Image(io.BytesIO(bar_png), width=6.2 * inch, height=2.6 * inch))
        story.append(Spacer(1, 0.2 * inch))

    pie_png = _pie_chart_png(process_breakdown, "Completions by Process Type")
    if pie_png:
        story.append(Image(io.BytesIO(pie_png), width=3.4 * inch, height=3.0 * inch))
        story.append(Spacer(1, 0.25 * inch))

    story.append(Paragraph("File Activity Detail", styles["Heading3"]))
    table_data = [["File Name", "Process Type", "Assigned To", "Status", "Assigned TS", "Completion TS"]]
    for row in rows:
        table_data.append(
            [
                row["file_name"],
                row["process_type"],
                row["assigned_to"],
                row["status"],
                row["assigned_ts"].strftime("%Y-%m-%d %H:%M") if row["assigned_ts"] else "-",
                row["completion_ts"].strftime("%Y-%m-%d %H:%M") if row["completion_ts"] else "-",
            ]
        )
    if len(table_data) == 1:
        table_data.append(["No activity in this range", "", "", "", "", ""])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C72B0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
            ]
        )
    )
    story.append(table)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, title="Completions Report")
    doc.build(story)
    return buf.getvalue(), label
